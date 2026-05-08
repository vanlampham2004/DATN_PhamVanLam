import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HEADERS = [
    "Test Case ID",
    "Description",
    "Expected Result",
    "Output",
    "Highlight"
]


def save_to_excel(raw_text, file_path):
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # lưu raw AI để kiểm tra AI trả gì
    raw_path = file_path.replace(".xlsx", "_raw.txt")
    with open(raw_path, "w", encoding="utf-8") as f:
        f.write(raw_text or "")

    rows = []

    for line in (raw_text or "").splitlines():
        line = line.strip().strip("|").strip()

        if not line:
            continue

        if "|" not in line:
            continue

        parts = [p.strip() for p in line.split("|")]

        if len(parts) != 5:
            continue

        tc_id = parts[0].strip()

        # bỏ header
        if tc_id.lower() == "test case id":
            continue

        # chỉ nhận TC001, TC002...
        if not re.match(r"^TC\d{3}$", tc_id):
            continue

        rows.append(parts)

    wb = Workbook()
    ws = wb.active
    ws.title = "UI Test Cases"

    ws.append(HEADERS)

    # nếu AI không sinh đúng format thì ghi thông báo, không hard-code testcase
    if not rows:
        ws.append([
            "AI_ERROR",
            "AI did not return valid test cases",
            "Ollama should return rows in required format",
            "Check raw AI response file",
            "UI_NOT_FOUND"
        ])
    else:
        for row in rows:
            ws.append(row)

    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col_letter].width = min(max_len + 5, 70)

    wb.save(file_path)